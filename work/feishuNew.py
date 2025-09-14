import json
import os
import logging
import threading
import time
import random
import shutil
from datetime import datetime
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from queue import Queue

from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException,
    ElementClickInterceptedException, StaleElementReferenceException,
    WebDriverException
)
from webdriver_manager.chrome import ChromeDriverManager

# -------------------------- 全局配置（效率优先）--------------------------
# 线程数：建议设为 CPU核心数×1.5~2（如8核设12，16核设20，避免内存溢出）
MAX_THREADS = 12
TIMEOUT_BASE = 20  # 基础超时时间（缩短，表格文件小）
TIMEOUT_DOWNLOAD_MULTI = 30  # 多文件下载超时（比单文件稍长）
RETRY_TIMES = 1  # 失败重试次数
LOG_LEVEL = logging.INFO
EXCEL_PATH = "1111.xlsx"  # 修复原Excel文件名引号问题
BASE_DOWNLOAD_DIR = os.path.abspath("feishu_downloads_multi_fast")
LOGIN_TEMPLATE_DIR = os.path.abspath("feishu_login_template")
LOGIN_POOL_DIR = os.path.abspath("feishu_login_pool")
# -----------------------------------------------------------------------------

log_lock = Lock()
report_lock = Lock()
global_results = []
login_dir_queue = Queue(maxsize=MAX_THREADS)
# 任务队列：存储待处理的链接（索引+链接），线程从队列取任务（更灵活）
task_queue = Queue()
logger = None


# -------------------------- 基础工具函数（全链路优化）--------------------------
def init_logger():
    global logger
    if logger is None:
        logger = logging.getLogger()
        logger.setLevel(LOG_LEVEL)
        formatter = logging.Formatter('%(asctime)s - %(threadName)s - %(levelname)s - %(message)s')
        # 日志按小时分割（避免5000条记录过大）
        file_handler = logging.FileHandler(
            f"feishu_download_{datetime.now().strftime('%Y%m%d_%H')}.log",
            encoding='utf-8'
        )
        file_handler.setFormatter(formatter)
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(logging.Formatter('%(threadName)s - %(levelname)s - %(message)s'))
        logger.addHandler(file_handler)
        logger.addHandler(console_handler)
    return logger

logger = init_logger()

def get_hyperlinks_from_excel():
    """从Excel的AC列提取超链接"""
    hyperlinks = []
    try:
        wb = load_workbook(EXCEL_PATH)
        sheet = wb.active  # 假设操作第一个工作表
        # AC列的索引：Excel列AC对应0-based索引27（A=0，B=1...AC=27）
        AC_COLUMN = 0

        # 遍历行（跳过表头，从第2行开始）
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            cell = row[AC_COLUMN]
            if not cell.value:
                continue
            cell_value = cell.value.split("\"")[1].replace("\'", "")
            # if cell.hyperlink and cell.hyperlink.target:
            #     href = cell.hyperlink.target
                # 处理相对路径，拼接基础URL
            hyperlinks.append(cell_value)
        return hyperlinks
    except Exception as e:
        print(f"读取Excel失败: {e}")
        return []


def create_login_template():
    """预登录生成模板目录（仅1次）"""
    if os.path.exists(LOGIN_TEMPLATE_DIR):
        logger.info(f"登录模板已存在：{LOGIN_TEMPLATE_DIR}，跳过预登录")
        return True

    logger.info("=" * 50)
    logger.info("请在弹出浏览器中登录飞书（登录后自动关闭）")
    logger.info("=" * 50)

    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(f"user-data-dir={LOGIN_TEMPLATE_DIR}")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)

    try:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        driver.get("https://www.feishu.cn/")

        # 等待登录（120秒超时，每10秒提示一次）
        login_wait_time = 120
        for i in range(login_wait_time // 10):
            time.sleep(10)
            login_cookies = [c for c in driver.get_cookies() if c["name"].startswith(("passport_", "feishu_"))]
            if login_cookies:
                logger.info("登录成功，模板目录生成完成")
                driver.quit()
                return True
            remaining_time = login_wait_time - (i + 1) * 10
            logger.info(f"等待登录... 剩余{remaining_time}秒")

        driver.quit()
        logger.error("预登录超时，删除无效目录")
        if os.path.exists(LOGIN_TEMPLATE_DIR):
            shutil.rmtree(LOGIN_TEMPLATE_DIR)
        return False
    except Exception as e:
        logger.error(f"预登录失败: {str(e)}")
        if os.path.exists(LOGIN_TEMPLATE_DIR):
            shutil.rmtree(LOGIN_TEMPLATE_DIR)
        return False


def init_login_pool():
    """初始化登录资源池（仅MAX_THREADS个目录，1次复制）"""
    if os.path.exists(LOGIN_POOL_DIR):
        shutil.rmtree(LOGIN_POOL_DIR)
    os.makedirs(LOGIN_POOL_DIR, exist_ok=True)
    logger.info(f"初始化登录资源池：创建 {MAX_THREADS} 个固定目录")

    for i in range(1, MAX_THREADS + 1):
        pool_dir = os.path.join(LOGIN_POOL_DIR, f"login_thread_{i}")
        try:
            shutil.copytree(LOGIN_TEMPLATE_DIR, pool_dir)
            login_dir_queue.put(pool_dir)
            logger.debug(f"资源池添加目录：{os.path.basename(pool_dir)}")
        except Exception as e:
            logger.error(f"创建资源池目录 {i} 失败: {str(e)}")
            return False
    return True


def init_reusable_driver(login_dir):
    """初始化可复用浏览器（1个线程1个Driver，处理所有链接）"""
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(f"user-data-dir={login_dir}")
    # 无头模式极致优化：禁用所有非必要组件
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-images")  # 禁用图片（表格下载无需图片）
    chrome_options.add_argument("--disable-stylesheets")  # 禁用CSS
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")  # 解决内存不足
    chrome_options.add_argument("--fast-start")  # 加速浏览器启动
    chrome_options.add_argument("--disable-background-networking")  # 禁用后台网络请求
    chrome_options.add_argument("--disable-popup-blocking")  # 禁用弹窗阻止（避免影响下载）

    # 下载配置：多文件自动下载
    prefs = {
        "download.default_directory": "",  # 后续动态设置，这里留空
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": False,  # 禁用安全检查，加速下载
        "profile.default_content_setting_values.automatic_downloads": 1,  # 允许多文件下载
        "plugins.always_open_pdf_externally": True,
        "download.improve_download_behavior": 1  # Chrome加速下载配置
    }
    chrome_options.add_experimental_option("prefs", prefs)

    # 反检测+性能优化
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)
    chrome_options.add_argument("--disable-infobars")

    try:
        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=chrome_options
        )
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        driver.set_page_load_timeout(TIMEOUT_BASE * 1.5)  # 缩短页面加载超时
        driver.set_script_timeout(TIMEOUT_BASE)

        # 验证登录状态
        driver.get("https://www.feishu.cn/")
        WebDriverWait(driver, TIMEOUT_BASE).until(
            lambda d: any(c["name"].startswith(("passport_", "feishu_")) for c in d.get_cookies())
        )
        logger.debug(f"浏览器初始化成功（复用目录：{os.path.basename(login_dir)}）")
        return driver
    except Exception as e:
        logger.error(f"浏览器初始化失败: {str(e)}")
        return None


def set_download_dir_cdp(driver, download_dir):
    """动态设置下载目录（支持多文件）"""
    download_dir = os.path.abspath(download_dir)
    os.makedirs(download_dir, exist_ok=True)
    driver.execute_cdp_cmd(
        "Page.setDownloadBehavior",
        {
            "behavior": "allow",
            "downloadPath": download_dir,
            "eventsEnabled": True  # 启用下载事件监听（多文件需要）
        }
    )
    return download_dir


def wait_for_multi_download(download_dir, initial_files, timeout=TIMEOUT_DOWNLOAD_MULTI):
    """等待多文件下载完成：对比初始文件列表，无临时文件且有新文件"""
    start_time = time.time()
    while time.time() - start_time < timeout:
        # 1. 检查是否有临时下载文件
        temp_files = [f for f in os.listdir(download_dir) if f.endswith((".crdownload", ".part", ".tmp"))]
        if temp_files:
            time.sleep(0.3)  # 缩短轮询间隔，提升响应速度
            continue
        # 2. 检查是否有新文件生成（对比初始列表）
        current_files = set(os.listdir(download_dir))
        new_files = current_files - initial_files
        if new_files:
            logger.debug(f"多文件下载完成：{len(new_files)}个新文件")
            return True, list(new_files)
        time.sleep(0.5)
    logger.warning(f"多文件下载超时（{timeout}秒）：{download_dir}")
    return False, []


def find_all_download_buttons(driver):
    """修复：返回所有有效下载按钮（支持多文件）"""
    # 扩展选择器，覆盖飞书常见下载按钮场景（如“导出”“下载全部”）
    selectors = [
        "//button[contains(text(), '下载') or contains(text(), '导出') or contains(@class, 'download') or contains(@class, 'export')]",
        "//a[contains(text(), '下载') or contains(text(), '导出') or contains(@class, 'download') or contains(@class, 'export')]",
        "//div[contains(text(), '下载') or contains(text(), '导出') and (@role='button' or contains(@class, 'button'))]",
        "//*[@aria-label='下载' or @aria-label='导出' or @title='下载' or @title='导出']",
        "//*[contains(@data-icon, 'download') or contains(@data-icon, 'export')]/parent::button"
    ]

    all_buttons = []
    for selector in selectors:
        try:
            # 显式等待按钮加载，避免漏找
            buttons = WebDriverWait(driver, TIMEOUT_BASE / 2).until(
                EC.presence_of_all_elements_located((By.XPATH, selector))
            )
            # 过滤可见且有效（非隐藏/禁用）的按钮
            valid_buttons = [
                btn for btn in buttons
                if btn.is_displayed() and btn.is_enabled() and btn.size['width'] > 10 and btn.size['height'] > 10
            ]
            all_buttons.extend(valid_buttons)
        except TimeoutException:
            continue
        except Exception as e:
            logger.debug(f"选择器 {selector} 查找异常: {str(e)}")

    # 去重：通过位置+文本去重（避免同一按钮被多个选择器匹配）
    unique_buttons = []
    seen = set()
    for btn in all_buttons:
        try:
            btn_text = btn.text.strip() or "无文本按钮"
            btn_key = (btn.location['x'], btn.location['y'], btn_text)
            if btn_key not in seen:
                seen.add(btn_key)
                unique_buttons.append(btn)
        except StaleElementReferenceException:
            continue

    logger.debug(f"找到 {len(unique_buttons)} 个有效下载按钮")
    return unique_buttons


# -------------------------- 核心优化：线程复用浏览器+多文件下载 --------------------------
def thread_worker(thread_id):
    """线程工作函数：1个线程复用1个浏览器，处理多个链接（核心效率优化）"""
    thread_name = f"Worker-{thread_id}"
    driver = None
    login_dir = None

    try:
        # 1. 从资源池获取登录目录，初始化浏览器（仅1次/线程）
        login_dir = login_dir_queue.get(timeout=30)
        if not login_dir:
            logger.error(f"{thread_name} 无法获取登录目录，退出线程")
            return

        driver = init_reusable_driver(login_dir)
        if not driver:
            logger.error(f"{thread_name} 浏览器初始化失败，退出线程")
            return

        # 2. 循环处理任务队列中的链接（复用浏览器）
        while not task_queue.empty():
            try:
                # 获取任务（超时10秒，避免空等）
                link_idx, link = task_queue.get(timeout=10)
                link_dir = os.path.join(BASE_DOWNLOAD_DIR, f"link_{link_idx}")
                result = {
                    "link_idx": link_idx,
                    "link": link,
                    "success": False,
                    "message": "",
                    "file_count": 0,
                    "files": [],
                    "save_dir": link_dir,
                    "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }

                # 3. 处理单个链接（多文件下载）
                try:
                    # 3.1 动态设置下载目录
                    set_download_dir_cdp(driver, link_dir)
                    # 3.2 记录初始文件（避免统计旧文件）
                    initial_files = set(os.listdir(link_dir)) if os.path.exists(link_dir) else set()

                    # 3.3 加载链接（打开新标签页，避免清理前页面）
                    driver.execute_script(f"window.open('{link}', '_blank');")
                    # 切换到新标签页（最后一个标签页）
                    driver.switch_to.window(driver.window_handles[-1])
                    logger.info(f"{thread_name} 加载链接[{link_idx}]: {link[:50]}...")

                    # 3.4 验证登录与页面加载
                    WebDriverWait(driver, TIMEOUT_BASE).until(
                        EC.presence_of_element_located((By.TAG_NAME, "body"))
                    )
                    if "登录" in driver.title or "未授权" in driver.page_source:
                        raise Exception("登录状态失效，需重新生成模板")

                    # 3.5 查找所有下载按钮，逐个点击
                    buttons = find_all_download_buttons(driver)
                    if not buttons:
                        raise Exception("未找到任何有效下载按钮")

                    # 3.6 遍历按钮点击（支持多文件）
                    clicked_buttons = 0
                    for btn_idx, btn in enumerate(buttons, 1):
                        try:
                            # 滚动到按钮可见位置（显式等待滚动完成）
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'instant'});", btn)
                            WebDriverWait(driver, TIMEOUT_BASE / 2).until(
                                EC.visibility_of(btn)
                            )
                            # JS点击（避免遮挡，比原生点击更稳定）
                            driver.execute_script("arguments[0].click();", btn)
                            clicked_buttons += 1
                            logger.debug(f"{thread_name} 点击第{btn_idx}个按钮（共{len(buttons)}个）")
                            time.sleep(0.3)  # 短等待，避免按钮连续点击冲突
                        except Exception as e:
                            logger.warning(f"{thread_name} 点击第{btn_idx}个按钮失败: {str(e)}")
                            continue

                    if clicked_buttons == 0:
                        raise Exception("所有按钮点击失败")

                    # 3.7 等待所有文件下载完成
                    download_success, new_files = wait_for_multi_download(link_dir, initial_files)
                    if download_success and new_files:
                        result.update({
                            "success": True,
                            "message": f"成功点击{clicked_buttons}个按钮，下载{len(new_files)}个文件",
                            "file_count": len(new_files),
                            "files": new_files
                        })
                        logger.info(f"{thread_name} 成功[{link_idx}]: {len(new_files)}个文件 → {link_dir}")
                    else:
                        raise Exception("点击按钮后无文件生成或下载超时")

                except Exception as e:
                    result["message"] = str(e)
                    logger.warning(f"{thread_name} 处理链接[{link_idx}]失败: {str(e)}")

                finally:
                    # 3.8 清理：关闭当前标签页，切换回初始标签页（避免内存泄漏）
                    if len(driver.window_handles) > 1:
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])
                    # 3.9 线程安全写入结果
                    with report_lock:
                        global_results.append(result)
                    # 3.10 标记任务完成（避免队列阻塞）
                    task_queue.task_done()

            except Exception as e:
                # 单个任务异常不影响线程继续处理下一个
                logger.error(f"{thread_name} 任务处理异常: {str(e)}")
                if not task_queue.empty():
                    task_queue.task_done()
                continue

    except Exception as e:
        logger.error(f"{thread_name} 线程全局异常: {str(e)}")

    finally:
        # 4. 线程结束：关闭浏览器+归还登录目录
        if driver:
            try:
                driver.quit()
                logger.info(f"{thread_name} 浏览器已关闭")
            except Exception as e:
                logger.error(f"{thread_name} 浏览器关闭失败: {str(e)}")
        if login_dir and os.path.exists(login_dir):
            return_login_dir_to_pool(login_dir)
        logger.info(f"{thread_name} 线程退出")


def get_login_dir_from_pool():
    """从资源池获取登录目录（线程安全）"""
    try:
        login_dir = login_dir_queue.get(timeout=30)
        logger.debug(f"线程 {threading.current_thread().name} 获取登录目录：{os.path.basename(login_dir)}")
        return login_dir
    except Exception as e:
        logger.error(f"获取登录目录失败: {str(e)}")
        return None


def return_login_dir_to_pool(login_dir):
    """归还登录目录到资源池（线程安全）"""
    if login_dir and os.path.exists(login_dir) and not login_dir_queue.full():
        login_dir_queue.put(login_dir)
        logger.debug(f"归还登录目录：{os.path.basename(login_dir)}")


# -------------------------- 批量调度（任务队列+线程池）--------------------------
def batch_process_fast_multi():
    # 1. 预登录+初始化资源池（仅1次）
    if not create_login_template():
        logger.error("预登录失败，退出")
        return
    if not init_login_pool():
        logger.error("登录资源池初始化失败，退出")
        return

    # 2. 读取链接+初始化任务队列
    os.makedirs(BASE_DOWNLOAD_DIR, exist_ok=True)
    links = get_hyperlinks_from_excel()
    if not links:
        logger.error("无有效链接，退出")
        return
    total_links = len(links)
    # 填充任务队列（链接索引+链接）
    for idx, link in enumerate(links, 1):
        task_queue.put((idx, link))
    logger.info(f"\n批量处理启动：{total_links}个链接，{MAX_THREADS}个线程（浏览器复用+多文件下载）")

    # 3. 启动线程池（固定线程数，复用浏览器）
    start_time = time.time()
    with ThreadPoolExecutor(max_workers=MAX_THREADS, thread_name_prefix="FeishuWorker") as executor:
        # 提交线程任务（每个线程对应1个worker）
        for thread_id in range(1, MAX_THREADS + 1):
            executor.submit(thread_worker, thread_id)

        # 等待所有任务完成
        task_queue.join()
        logger.info(f"所有任务处理完成，开始生成报告")

    # 4. 生成报告+清理资源池
    generate_final_report(global_results, total_links)
    clean_login_pool()
    total_time = time.time() - start_time
    logger.info(f"批量处理全部完成！总耗时：{total_time:.0f}秒（平均{total_time/total_links:.2f}秒/链接）")


def generate_final_report(results, total_links):
    """生成详细报告，包含多文件下载统计"""
    success_count = sum(1 for r in results if r["success"])
    total_files = sum(r["file_count"] for r in results)
    fail_links = [r for r in results if not r["success"]]

    # 按链接索引排序（便于查看）
    results_sorted = sorted(results, key=lambda x: x["link_idx"])

    # 保存JSON报告（便于后续重跑失败链接）
    report_json = os.path.join(BASE_DOWNLOAD_DIR, "batch_report_multi.json")
    with open(report_json, 'w', encoding='utf-8') as f:
        json.dump(results_sorted, f, ensure_ascii=False, indent=2)

    # 保存文本报告（关键信息一目了然）
    report_txt = os.path.join(BASE_DOWNLOAD_DIR, "batch_report_multi.txt")
    with open(report_txt, 'w', encoding='utf-8') as f:
        f.write(f"飞书多文件批量下载报告 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 60 + "\n")
        f.write(f"总链接数：{total_links}\n")
        f.write(f"成功链接数：{success_count} ({(success_count/total_links)*100:.1f}%)\n")
        f.write(f"失败链接数：{len(fail_links)} ({(len(fail_links)/total_links)*100:.1f}%)\n")
        f.write(f"总下载文件数：{total_files}\n")
        f.write(f"平均每个成功链接文件数：{total_files/success_count:.1f}（若成功数>0）\n")
        f.write(f"基础保存目录：{BASE_DOWNLOAD_DIR}\n")
        f.write("=" * 60 + "\n")

        # 输出前20个失败链接（便于快速定位问题）
        if fail_links:
            f.write(f"\n失败链接示例（前20个）：\n")
            for fail in sorted(fail_links[:20], key=lambda x: x["link_idx"]):
                f.write(f"  索引{fail['link_idx']}：{fail['link'][:60]}... | 原因：{fail['message'][:60]}...\n")
            if len(fail_links) > 20:
                f.write(f"  ...（剩余{len(fail_links)-20}个失败链接详见JSON报告）\n")

    logger.info(f"报告已保存至：{report_txt} 和 {report_json}")
    logger.info(f"核心统计：{success_count}/{total_links}链接成功，共下载{total_files}个文件")


def clean_login_pool():
    """任务结束后统一清理资源池（仅1次）"""
    if os.path.exists(LOGIN_POOL_DIR):
        try:
            shutil.rmtree(LOGIN_POOL_DIR)
            logger.info(f"登录资源池已清理：{LOGIN_POOL_DIR}")
        except Exception as e:
            logger.error(f"清理资源池失败: {str(e)}")


# -------------------------- 入口函数 --------------------------
if __name__ == "__main__":
    # 预处理：检查Excel文件（修复原代码引号问题）
    if not os.path.exists(EXCEL_PATH):
        logger.error(f"Excel文件不存在：{EXCEL_PATH}")
        exit(1)

    try:
        batch_process_fast_multi()
    except KeyboardInterrupt:
        logger.error("任务被手动中断")
        # 清空任务队列+清理资源池
        while not task_queue.empty():
            task_queue.get()
            task_queue.task_done()
        clean_login_pool()
    except Exception as e:
        logger.error(f"全局异常: {str(e)}")
        clean_login_pool()