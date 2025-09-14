import json

from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementClickInterceptedException, \
    StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
import time
import os
import logging
from datetime import datetime

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("feishu_download.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
# EXCEL_PATH = "1111.xlsx"
EXCEL_PATH = "1111.xlsx"

def get_hyperlinks_from_excel():
    """从Excel的AC列提取超链接"""
    hyperlinks = []
    try:
        wb = load_workbook(EXCEL_PATH)
        sheet = wb.active  # 假设操作第一个工作表
        # AC列的索引：Excel列AC对应0-based索引27（A=0，B=1...AC=27）
        AC_COLUMN = 0

        # 遍历行（跳过表头，从第2行开始）
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            cell = row[AC_COLUMN]
            if not cell.value:
                continue
            cell_value = cell.value.split("\"")[1].replace("\'", "")
            # if cell.hyperlink and cell.hyperlink.target:
            #     href = cell.hyperlink.target
                # 处理相对路径，拼接基础URL
            hyperlinks.append(cell_value)
        return hyperlinks
    except Exception as e:
        print(f"读取Excel失败: {e}")
        return []


def init_driver(base_download_dir):
    """初始化WebDriver并配置基础选项"""
    chrome_options = webdriver.ChromeOptions()

    # 确保基础下载目录存在并获取绝对路径
    base_download_dir = os.path.abspath(base_download_dir)
    os.makedirs(base_download_dir, exist_ok=True)

    # 禁用用户数据目录，避免默认设置覆盖
    chrome_options.add_argument("--no-user-data-dir")

    # 基础下载设置
    prefs = {
        "download.default_directory": base_download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "profile.default_content_setting_values.automatic_downloads": 1,
        "profile.default_content_settings.popups": 0,
        "plugins.always_open_pdf_externally": True,
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False
    }
    chrome_options.add_experimental_option("prefs", prefs)

    # 反检测设置
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--start-maximized")

    # 初始化驱动
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=chrome_options
    )

    # 进一步反检测
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    logging.info(f"初始化浏览器，基础下载路径: {base_download_dir}")
    return driver


def set_download_directory(driver, download_dir):
    """动态设置浏览器下载路径"""
    download_dir = os.path.abspath(download_dir)
    os.makedirs(download_dir, exist_ok=True)

    # 通过Chrome DevTools协议设置下载路径
    driver.execute_cdp_cmd(
        "Page.setDownloadBehavior",
        {
            "behavior": "allow",
            "downloadPath": download_dir,
            "eventsEnabled": True
        }
    )
    logging.info(f"已设置当前下载路径: {download_dir}")
    time.sleep(2)  # 等待设置生效
    return download_dir


def switch_to_possible_iframe(driver):
    """切换到可能包含下载按钮的iframe"""
    try:
        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        logging.info(f"发现 {len(iframes)} 个iframe，尝试切换...")

        for i, iframe in enumerate(iframes):
            try:
                driver.switch_to.frame(iframe)
                logging.info(f"已切换到第 {i + 1} 个iframe")
                return True
            except Exception as e:
                logging.debug(f"切换到第 {i + 1} 个iframe失败: {str(e)}")
                continue

        driver.switch_to.default_content()
        return False
    except Exception as e:
        logging.warning(f"处理iframe时出错: {str(e)}")
        driver.switch_to.default_content()
        return False


def wait_for_downloads(download_dir, timeout=120):
    """等待下载完成"""
    start_time = time.time()
    while time.time() - start_time < timeout:
        temp_files = [
            f for f in os.listdir(download_dir)
            if f.endswith(('.crdownload', '.part', '.download', '.tmp'))
        ]
        if not temp_files:
            return True
        time.sleep(2)
    return False


def find_download_buttons(driver):
    """多策略查找下载按钮"""
    download_buttons = []

    # 文本包含"下载"的元素
    text_based_selectors = [
        "//button[contains(text(), '下载')]",
        "//span[contains(text(), '下载')]/parent::button",
        "//a[contains(text(), '下载')]",
        "//div[contains(text(), '下载') and (@role='button' or contains(@class, 'button'))]",
        "//li[contains(text(), '下载')]",
        "//menu[contains(text(), '下载')]"
    ]

    # 包含下载相关图标的元素
    icon_based_selectors = [
        "//button[contains(@class, 'download') or contains(@class, 'export')]",
        "//a[contains(@class, 'download') or contains(@class, 'export')]",
        "//div[contains(@class, 'download') or contains(@class, 'export') and (@role='button' or contains(@class, 'button'))]",
        "//*[contains(@class, 'icon-download') or contains(@class, 'icon-export')]/parent::button",
        "//*[contains(@data-icon, 'download') or contains(@data-icon, 'export')]/parent::button"
    ]

    # 包含下载相关属性的元素
    attribute_based_selectors = [
        "//*[@data-action='download']",
        "//*[@data-type='download']",
        "//*[@aria-label='下载']",
        "//*[@title='下载']"
    ]

    all_selectors = text_based_selectors + icon_based_selectors + attribute_based_selectors

    for selector in all_selectors:
        try:
            elements = driver.find_elements(By.XPATH, selector)
            for elem in elements:
                if elem.is_displayed() and elem.size['width'] > 0 and elem.size['height'] > 0:
                    download_buttons.append(elem)
        except Exception as e:
            logging.debug(f"选择器 {selector} 查找失败: {str(e)}")
            continue

    # 去重按钮
    unique_buttons = []
    seen = set()
    for btn in download_buttons:
        try:
            key = (btn.location['x'], btn.location['y'], btn.size['width'], btn.size['height'])
            if key not in seen:
                seen.add(key)
                unique_buttons.append(btn)
        except StaleElementReferenceException:
            continue

    return unique_buttons


def download_multi_files(driver, url, download_dir):
    """下载单个页面上的多个文件"""
    # 关键修复：动态设置当前链接的下载路径
    download_dir = set_download_directory(driver, download_dir)

    file_results = []
    try:
        driver.get(url)
        logging.info(f"正在加载页面: {url}")

        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        time.sleep(5)

        iframe_switched = switch_to_possible_iframe(driver)
        initial_files = set(os.listdir(download_dir))

        # 多次尝试查找按钮
        download_buttons = []
        for attempt in range(3):
            buttons = find_download_buttons(driver)
            if buttons:
                download_buttons = buttons
                break
            logging.info(f"第 {attempt + 1} 次尝试查找下载按钮...")
            time.sleep(3)

        if not download_buttons and iframe_switched:
            driver.switch_to.default_content()
            logging.info("切回主文档再次尝试查找下载按钮...")
            download_buttons = find_download_buttons(driver)

        if not download_buttons:
            logging.info("尝试滚动页面以触发内容加载...")
            for i in range(3):
                driver.execute_script(f"window.scrollTo(0, {i * 500});")
                time.sleep(2)
                download_buttons = find_download_buttons(driver)
                if download_buttons:
                    break

        if not download_buttons:
            logging.warning("未找到下载按钮，是否需要手动操作？")
            user_input = input("请在浏览器中手动找到下载按钮位置，按Enter继续（输入q退出）: ").strip().lower()
            if user_input == 'q':
                return False, "用户选择退出", file_results

            time.sleep(15)
            post_files = set(os.listdir(download_dir))
            new_files = post_files - initial_files
            if new_files:
                wait_for_downloads(download_dir)
                file_results.append({
                    "button": "手动下载",
                    "success": True,
                    "files": list(new_files),
                    "message": "手动下载成功"
                })
                return True, "手动下载成功", file_results
            else:
                return False, "手动操作未检测到新文件", file_results

        logging.info(f"找到 {len(download_buttons)} 个下载按钮")

        # 逐个处理下载按钮
        for i, button in enumerate(download_buttons, 1):
            try:
                pre_click_files = set(os.listdir(download_dir))

                # 移动到按钮位置
                actions = ActionChains(driver)
                actions.move_to_element(button).perform()
                time.sleep(1)

                # 滚动到可见位置
                driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", button)
                time.sleep(2)

                # 尝试多种点击方式
                clicked = False
                for click_attempt in range(3):
                    try:
                        if click_attempt == 0:
                            button.click()  # 常规点击
                        else:
                            driver.execute_script("arguments[0].click();", button)
                        clicked = True
                        break
                    except Exception as e:
                        logging.debug(f"第 {click_attempt + 1} 次点击尝试失败: {str(e)}")
                        time.sleep(1)

                if not clicked:
                    logging.warning(f"第 {i} 个按钮点击失败")
                    file_results.append({
                        "button": f"第{i}个按钮",
                        "success": False,
                        "files": [],
                        "message": "无法点击按钮"
                    })
                    continue

                button_text = button.text.strip() or f"第{i}个按钮"
                logging.info(f"已点击: {button_text}")
                time.sleep(5)

                # 检查新文件
                post_click_files = set(os.listdir(download_dir))
                new_files = post_click_files - pre_click_files - initial_files

                # 等待下载完成
                if new_files or wait_for_downloads(download_dir):
                    final_files = set(os.listdir(download_dir)) - pre_click_files - initial_files
                    if final_files:
                        logging.info(f"{button_text} 下载成功，文件: {len(final_files)} 个")
                        file_results.append({
                            "button": button_text,
                            "success": True,
                            "files": list(final_files),
                            "message": f"下载成功"
                        })
                    else:
                        logging.warning(f"{button_text} 未生成新文件")
                        file_results.append({
                            "button": button_text,
                            "success": False,
                            "files": [],
                            "message": "未生成新文件"
                        })
                else:
                    logging.warning(f"{button_text} 下载超时")
                    file_results.append({
                        "button": button_text,
                        "success": False,
                        "files": list(new_files),
                        "message": "下载超时"
                    })

                time.sleep(5)

            except Exception as e:
                err_msg = f"处理第{i}个按钮时出错: {str(e)}"
                logging.error(err_msg)
                file_results.append({
                    "button": f"第{i}个按钮",
                    "success": False,
                    "files": [],
                    "message": err_msg
                })

        if iframe_switched:
            driver.switch_to.default_content()

        any_success = any(res["success"] for res in file_results)
        return any_success, f"处理完成，共 {len(download_buttons)} 个下载按钮", file_results

    except Exception as e:
        err_msg = f"处理页面时出错: {str(e)}"
        return False, err_msg, file_results


def batch_download_files(download_dir=None):
    """批量下载主函数"""
    if not download_dir:
        download_dir = os.path.join(os.getcwd(), "feishu_downloads")
    # 确保基础下载目录使用绝对路径
    base_download_dir = os.path.abspath(download_dir)
    os.makedirs(base_download_dir, exist_ok=True)
    logging.info(f"基础下载目录: {base_download_dir}")

    links = get_hyperlinks_from_excel()

    if not links:
        logging.warning("链接文件为空，没有需要下载的内容")
        return

    logging.info(f"共发现 {len(links)} 个链接，开始批量下载...")

    driver = init_driver(base_download_dir)
    results = []

    try:
        logging.info("请确认飞书登录状态，5秒后开始下载...")
        time.sleep(5)

        for i, link in enumerate(links, 1):
            logging.info(f"\n===== 处理第 {i}/{len(links)} 个链接 =====")

            # 为每个链接创建独立目录
            link_dir = os.path.join(base_download_dir, f"link_{i}")
            os.makedirs(link_dir, exist_ok=True)
            link_dir = os.path.abspath(link_dir)

            # 下载当前链接的文件
            success, msg, file_results = download_multi_files(driver, link, link_dir)

            total_files = len(file_results)
            success_files = sum(1 for res in file_results if res["success"])

            results.append({
                "link": link,
                "total_buttons": total_files,
                "success_buttons": success_files,
                "success": success,
                "message": msg,
                "file_results": file_results,
                "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "save_dir": link_dir
            })

            if success:
                logging.info(f"第 {i} 个链接处理完成，成功下载 {success_files}/{total_files} 个文件")
            else:
                logging.error(f"第 {i} 个链接处理失败: {msg}")

            time.sleep(5)

    except Exception as e:
        logging.error(f"批量处理发生错误: {str(e)}")
    finally:
        # 保存结果报告
        json_report = os.path.join(base_download_dir, "download_report.json")
        with open(json_report, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)

        report_file = os.path.join(base_download_dir, "download_report.txt")
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(f"飞书文件批量下载报告 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"总链接数: {len(links)}\n")
            f.write(f"成功处理链接数: {sum(1 for r in results if r['success'])}\n")
            f.write(f"总下载按钮数: {sum(r['total_buttons'] for r in results)}\n")
            f.write(f"成功下载数: {sum(r['success_buttons'] for r in results)}\n\n")

            f.write("详细结果:\n")
            for i, res in enumerate(results, 1):
                status = "成功" if res['success'] else "失败"
                f.write(f"{i}. 链接: {res['link']}\n")
                f.write(f"   状态: {status}\n")
                f.write(f"   时间: {res['time']}\n")
                f.write(f"   保存目录: {res['save_dir']}\n")
                f.write(f"   下载按钮: {res['success_buttons']}/{res['total_buttons']} 个成功\n")
                f.write(f"   信息: {res['message']}\n\n")

        driver.quit()
        logging.info(f"浏览器已关闭，报告已保存至: {report_file} 和 {json_report}")

    logging.info("批量下载任务完成")


if __name__ == "__main__":
    batch_download_files(None)