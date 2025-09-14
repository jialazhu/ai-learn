from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import os

# -------------------- 配置参数 --------------------
# 本地Excel文件路径（请替换为实际路径）
# EXCEL_PATH = "付款申请 (21年9月-25年1月 提报的打款申请.xlsx"
EXCEL_PATH = "1111.xlsx"
# 内部系统基础URL（若超链接为相对路径，需拼接；若为绝对路径可留空，或根据实际修改）
BASE_URL = "https://your.company.intranet"
# 文件保存目录
DOWNLOAD_DIR = "downloads"


def init_download_dir():
    """创建下载目录"""
    if not os.path.exists(DOWNLOAD_DIR):
        os.makedirs(DOWNLOAD_DIR)
        print(f"创建目录: {DOWNLOAD_DIR}")


def get_hyperlinks_from_excel():
    """从Excel的AC列提取超链接"""
    hyperlinks = []
    try:
        wb = load_workbook(EXCEL_PATH)
        sheet = wb.active  # 假设操作第一个工作表
        # AC列的索引：Excel列AC对应0-based索引27（A=0，B=1...AC=27）
        AC_COLUMN = 0

        # 遍历行（跳过表头，从第2行开始）
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            cell = row[AC_COLUMN]
            if not cell.value:
                continue
            cell_value = cell.value.split("\"")[1].replace("\'", "")
            # if cell.hyperlink and cell.hyperlink.target:
            #     href = cell.hyperlink.target
                # 处理相对路径，拼接基础URL
            if not cell_value.startswith(("http://", "https://")):
                cell_value = f"{BASE_URL}{cell_value}"
            hyperlinks.append(cell_value)
        return hyperlinks
    except Exception as e:
        print(f"读取Excel失败: {e}")
        return []

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    # 如有需要可添加Cookie信息
    "set-cookie": "passport_app_access_token=eyJhbGciOiJFUzI1NiIsInR5cCI6IkpXVCJ9.eyJleHAiOjE3NTYzMjg2NTcsInVuaXQiOiJldV9uYyIsInJhdyI6eyJtX2FjY2Vzc19pbmZvIjp7IjExOSI6eyJpYXQiOjE3NTYyODU0NTcsImFjY2VzcyI6dHJ1ZX19LCJzdW0iOiIyODFjMTBjOWYyMjRkMDg3YWRlZjdiNTc4ZDQzNzA5OTRhNWJjNDYxZjU5MzgzMmM4MTljMmIxZTJkMzhmYTY3In19.FKWHzrhQL9nMWQwZZv9vJUXdB_YNyq1-t3T6c4mdCrEluafMP7OSTiYJylDIduF-iY53_k0muoxEoJjOuOE4QQ; Path=/; Max-Age=43200; HttpOnly; Secure; SameSite=None"
}
def decode_content(content):
    encodings = ["utf-8", "GBK", "GB2312", "ISO-8859-1", "utf-16"]
    for encoding in encodings:
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return "无法解码"

def download_files_from_page(url):
    """访问页面，解析并下载所有“下载”链接的文件"""

    try:
        response = requests.get(url, headers=HEADERS, timeout=10)
        response.raise_for_status()  # 检查HTTP错误
        soup = BeautifulSoup(response.text, "html.parser")

        content_type = response.headers

        # 解析下载链接（假设“下载”文字是按钮标识，可根据实际页面调整）
        download_links = []
        for a_tag in soup.find_all("a", string="下载"):  # 若按钮有class，可改为soup.find_all("a", class_="download-btn")
            dl_href = a_tag.get("href")
            if dl_href:
                # 处理相对路径
                if not dl_href.startswith(("http://", "https://")):
                    dl_href = f"{BASE_URL}{dl_href}"
                download_links.append(dl_href)

        # 下载每个文件
        for dl_link in download_links:
            try:
                dl_response = requests.get(dl_link, headers=headers, stream=True, timeout=20)
                dl_response.raise_for_status()

                # 提取文件名（优先从响应头，其次从URL）
                filename = os.path.basename(dl_link)
                if "Content-Disposition" in dl_response.headers:
                    cd = dl_response.headers["Content-Disposition"]
                    filename = cd.split('filename=')[-1].strip('"')

                filepath = os.path.join(DOWNLOAD_DIR, filename)
                with open(filepath, "wb") as f:
                    for chunk in dl_response.iter_content(chunk_size=8192):
                        f.write(chunk)
                print(f"✅ 下载成功: {filename}")
            except Exception as e:
                print(f"❌ 下载失败 ({dl_link}): {e}")
    except Exception as e:
        print(f"❌ 访问页面失败 ({url}): {e}")


if __name__ == "__main__":
    init_download_dir()
    hyperlinks = get_hyperlinks_from_excel()
    if not hyperlinks:
        print("没有提取到有效超链接！")
    else:
        print(f"共提取 {len(hyperlinks)} 个超链接，开始处理...")
        for link in hyperlinks:
            print(f"\n处理页面: {link}")
            download_files_from_page(link)
    print("\n=== 任务完成 ===")